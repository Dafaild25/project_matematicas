from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse # Importar para enviar respuesta HTTP
from django.http import JsonResponse # Importar para enviar respuesta JSON
from django.contrib import messages  # Importar mensajes
from django.contrib.auth.models import User,Group # Importar user y grupos
from Aplicaciones.core.forms.user_form import * # Importar clase de registro
from ..models import * # Importar modelos
from ..decorators import admin_required
from datetime import datetime # Importar fecha y hora
from django.forms.utils import ErrorDict # Importar para presentar una vista mejor de los errores
from Aplicaciones.core.diccionarios.NombreCampos import Fields_names # Diccionario con los nombres de los campos formateados
# PLANTILLA EXCEL
import pandas as pd
from openpyxl import Workbook, load_workbook # Importar librería para crear archivos excel
import io
from openpyxl.styles import Font

# VISTA PRINCIPAL PARA LISTAR DOCENTES
@admin_required
def index(request):
    # Obtener todos los docentes excluyendo al usuario logueado
    docentes = Docentes.objects.exclude(fk_id_persona__fk_id_usuario=request.user) 
    return render(request, 'docente/index.html', {'docentes': docentes})

# METODO PARA DESCARGAR PLANTILLA EN EXCEL
@admin_required
def descargar_plantilla_docentes(request):
    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Docentes"
    # Encabezados ordenados y estilizados
    columnas = ['primer_nombre', 'segundo_nombre', 'primer_apellido', 'segundo_apellido','fecha_nacimiento', 'cedula', 'telefono','email']
    ws.append(columnas)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # Fila de ejemplo (valores como texto, conservando ceros)
    ejemplo = ['Juan', 'Carlos', 'Pérez', 'González','1985-10-22', '1729077856', '0912345678', 'juan.perez@example.com']
    ws.append(ejemplo)
    # Aplicar formato texto explícito a cédula y teléfono (columnas 6 y 7: F y G)
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = '@'
    # Ajuste de anchos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    # Preparar archivo para descarga
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_docentes.xlsx'
    return response

# METODO PARA EXPORTAR EL LISTADO DE DOCENTES A EXCEL
@admin_required
def exportar_docentes(request):
    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Docentes"
    # Encabezados
    columnas = [
        'Primer Nombre', 'Segundo Nombre', 'Primer Apellido', 'Segundo Apellido',
        'Fecha de Nacimiento', 'Cédula', 'Teléfono', 'Email', 'Estado'
    ]
    ws.append(columnas)
    # Aplicar negrita a los encabezados
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # Consultar todos los docentes
    docentes = Docentes.objects.all()
    # Agregar los datos al Excel
    for docente in docentes:
        persona = docente.fk_id_persona
        usuario = persona.fk_id_usuario
        fila = [
            usuario.first_name,  
            persona.per_segundo_nombre,       
            usuario.last_name,   
            persona.per_segundo_apellido,
            persona.per_fecha_nacimiento.strftime('%Y-%m-%d') if persona.per_fecha_nacimiento else '',
            persona.per_cedula,
            persona.per_telefono,
            usuario.email,
            'Activo' if docente.doc_estado else 'Inactivo'  # Estado (mostrar como texto)
        ]
        ws.append(fila)
    # Aplicar formato texto explícito a cédula y teléfono (columnas F y G)
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = '@'
    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    # Preparar archivo para descarga
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=docentes.xlsx'
    return response

# METODO PARA REGISTRAR DOCENTES DE FORMA MASIVA
@admin_required
def importar_docentes_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        try:
            # Obtener el archivo subido
            archivo_excel = request.FILES['archivo']
            # Verificar si el archivo es un archivo Excel válido
            df = pd.read_excel(archivo_excel, dtype=str)
            # Verificar si las columnas necesarias existen
            columnas_necesarias = ['primer_nombre', 'segundo_nombre', 'primer_apellido', 'segundo_apellido', 'fecha_nacimiento', 'cedula', 'telefono', 'email']
            for columna in columnas_necesarias:
                if columna not in df.columns:
                    return JsonResponse({'success': False, 'message': f'Falta la columna: {columna}'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error al leer el archivo: {str(e)}'})
        duplicados = [] # Array para almacenar cedulas o usuarios duplicados
        errores = [] # Array para almacenar errores de validación
        agregados = 0 # Variable para contar los docentes agregados
        # Recorrer las filas del DataFrame
        for index, row in df.iterrows():
            fila_num = index + 2  # Fila real en Excel
            primer_nombre = str(row.get('primer_nombre', '')).strip().title()
            segundo_nombre = str(row.get('segundo_nombre', '')).strip().title()
            primer_apellido = str(row.get('primer_apellido', '')).strip().title()
            segundo_apellido = str(row.get('segundo_apellido', '')).strip().title()
            fecha_nacimiento_str = str(row.get('fecha_nacimiento', '')).strip()
            # Verificar que exista una cedula sino envia Vacio, y quita puntos o comas
            cedula = str(row.get('cedula', '')).strip().replace('.', '').replace(',', '')
            telefono = str(row.get('telefono', '')).strip().replace('.', '').replace(',', '')
            # Si no hay email, agrega uno basado en el numero de cedula
            email = row.get('email', f"{cedula}@gmail.com").strip()
            # Verificar que la cedula no sea un digito
            if not cedula.isdigit():
                duplicados.append(f"Fila {index + 2} - Cédula inválida: {cedula}")
                continue  
            # Verificar si existe la cedula en otro usuario
            if(Personas.objects.filter(per_cedula=cedula).exists()):
                duplicados.append(f"Fila {index + 2} - Cédula duplicada: {cedula}")
                continue
            # Verificar si el username tiene el mismo numero de cedula
            if User.objects.filter(username=cedula).exists():
                duplicados.append(f"Fila {index + 2} - Username duplicado: {cedula}")
                continue
            # Procesar fecha de nacimiento
            fecha_nacimiento = None
            if fecha_nacimiento_str:
                try:
                    # Convertir en formato Y-M-D
                    fecha_nacimiento = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
                except ValueError: # Error de valor, el tipo de dato no corresponde con el formato
                    errores.append({
                        'fila': fila_num,
                        'errores_usuario': [],
                        'errores_persona': [f"Fecha de Nacimiento: Fecha inválida ({fecha_nacimiento_str})"]
                    })
                    continue
            # Validar User con formulario
            user_data = {
                'username': cedula,
                'first_name': primer_nombre,
                'last_name': primer_apellido,
                'email': email,
            }
            # Validar Persona con formulario
            persona_data = {
                'per_segundo_nombre': segundo_nombre,
                'per_segundo_apellido': segundo_apellido,
                'per_fecha_nacimiento': fecha_nacimiento,
                'cedula': cedula,
                'per_telefono': telefono,
            }
            formUsuario = UserUpdateForm(user_data)
            formPersona = PersonaCreateForm(persona_data)
            # Validar datos de User y Personas
            if(formUsuario.is_valid() and formPersona.is_valid()):
                nombre_grupo = 'Docentes' # Nombre del grupo
                grupo, created = Group.objects.get_or_create(name=nombre_grupo)
                # Datos para tabla User
                usuario = User.objects.create_user(
                    username = cedula,
                    first_name= primer_nombre,
                    last_name= primer_apellido,
                    email = email,
                    password=cedula
                )
                usuario.groups.add(grupo) # Agregar usuario al grupo  
                # Datos para tabla Personas
                persona = Personas.objects.create(
                    fk_id_usuario=usuario,
                    per_segundo_nombre=segundo_nombre,
                    per_segundo_apellido=segundo_apellido,
                    per_fecha_nacimiento=fecha_nacimiento,
                    per_cedula=cedula,
                    per_telefono=telefono
                )
                # Crear Docente
                docente = Docentes.objects.create(
                    fk_id_persona=persona
                )
                agregados += 1 # Contar de Docentes agregados
            else:
                errores_fila = {
                    'fila': fila_num,
                    'errores_usuario': [],
                    'errores_persona': []
                }
                if not formUsuario.is_valid():
                    for field, error_list in formUsuario.errors.items():
                        nombre_campo = Fields_names.get(field, field)
                        for error in error_list:
                            errores_fila['errores_usuario'].append(f"{nombre_campo}: {error}")
                if not formPersona.is_valid():
                    for field, error_list in formPersona.errors.items():
                        nombre_campo = Fields_names.get(field, field)
                        for error in error_list:
                            errores_fila['errores_persona'].append(f"{nombre_campo}: {error}")
                errores.append(errores_fila)
        # Clasificar el resultado
        if agregados > 0 and (len(duplicados) > 0 or len(errores) > 0):
            status = 'warning'
        elif agregados > 0:
            status = 'success'
        else:
            status = 'error'
        return JsonResponse({
            'success': True,  # Lo dejamos siempre en True para que JS lo maneje
            'status': status,
            'agregados': agregados,
            'duplicados': duplicados,
            'errores': errores
        })
    return JsonResponse({'error': False, 'message': 'Solicitud inválida'})

# VISTA PARA FORMULARIO DOCENTES
@admin_required
def create_docente(request):
    formUsuario = UserCreateForm()
    formPersona = PersonaCreateForm()
    return render(request, 'docente/Create.html',{
        'formUsuario': formUsuario,
        'formPersona': formPersona
    })

# METODO CREAR UN NUEVO DOCENTE
@admin_required
def nuevo_docente(request):
    formUsuario = UserCreateForm(request.POST or None) # Instanciar el formulario de usuario
    formPersona = PersonaCreateForm(request.POST or None) # Instanciar el formulario de persona
    if(request.method == 'POST'):
        cedula = request.POST.get('cedula', '').strip() # Obtener cedula del formulario
        cedula_existente = Personas.objects.filter(per_cedula=cedula).exists() # Verificar si existe la cedula
        if cedula_existente:
            formPersona.add_error('cedula', 'La cédula ya está registrada.') # Agregar error al formulario si la cedula ya existe
        if formUsuario.is_valid() and formPersona.is_valid():
            try:
                nombre_grupo = 'Docentes' # Nombre del grupo
                grupo, created = Group.objects.get_or_create(name=nombre_grupo)
                usuario = formUsuario.save() # Guardar el usuario
                usuario.first_name = usuario.first_name.title()
                usuario.last_name = usuario.last_name.title()
                usuario.save() # Guardar el usuario con los cambios
                usuario.groups.add(grupo) # Agregar usuario al grupo  
                # Datos para la tabla Personas
                persona = Personas.objects.create(
                    fk_id_usuario=usuario,
                    per_segundo_nombre=request.POST.get('per_segundo_nombre', '').title(),
                    per_segundo_apellido=request.POST.get('per_segundo_apellido', '').title(),
                    per_fecha_nacimiento=request.POST.get('per_fecha_nacimiento'),
                    per_cedula=request.POST.get('cedula'),
                    per_telefono=request.POST.get('per_telefono', '')
                )
                # Datos para la tabla Docentes
                docente = Docentes.objects.create( # Crear registro de tabla Docentes
                    fk_id_persona=persona,  # Relacionar con la persona creada
                    doc_fotografia= request.FILES.get('fotografia') if 'fotografia' in request.FILES else None,
                )
                messages.success(request,'Docente creado exitosamente.')
                return redirect('docente_index')
            except Exception as e:
                # Si ocurre un error, se renderiza de nuevo el formulario con los datos ingresados
                return render(request, 'docente/Create.html', {
                    'formUsuario': formUsuario,
                    'formPersona': formPersona
                })
    # Si el formulario no es válido, se renderiza de nuevo con los errores
    return render(request, 'docente/Create.html', {
        'formUsuario': formUsuario,
        'formPersona': formPersona
    })

# VISTA PARA EDITAR UN DOCENTE
@admin_required
def edit_docente(request, id_docen):
    docente = get_object_or_404(Docentes,pk=id_docen) # Obtener el docente por su ID
    persona = docente.fk_id_persona # Obtener la persona asociada al docente
    usuario = persona.fk_id_usuario # Obtener el usuario asociado a la persona
    # Crear los formularios precargados
    formUsuario = UserCreateForm(instance=usuario)
    initial_data = {
        'per_segundo_nombre': persona.per_segundo_nombre,
        'per_segundo_apellido': persona.per_segundo_apellido,
        'per_fecha_nacimiento': persona.per_fecha_nacimiento.strftime('%Y-%m-%d') if persona.per_fecha_nacimiento else '',
        'cedula': persona.per_cedula,
        'per_telefono': persona.per_telefono,
    }
    formPersona = PersonaCreateForm(initial=initial_data)
    contexto = {
        'formUsuario': formUsuario,
        'formPersona': formPersona,
        'docente': docente,
        'persona': persona,
    }
    return render(request,'docente/Edit.html', contexto) # Renderizar la plantilla de edición

# METODO PARA ACTUALIZAR UN DOCENTE
@admin_required
def actualizar_docente(request):
    if request.method == 'POST':
        # Obtener el ID del docente desde el formulario
        docen_id = request.POST['docen_id']
        docente = get_object_or_404(Docentes, pk=docen_id)  # Obtener el docente por su ID
        persona = docente.fk_id_persona  # Obtener la persona asociada al docente
        usuario = persona.fk_id_usuario  # Obtener el usuario asociado a la persona
        # Instanciar el formulario de usuario con los datos del POST
        formUsuario = UserUpdateForm(request.POST, instance=usuario)
        # Instanciar el formulario de persona con los datos del POST
        formPersona = PersonaCreateForm(request.POST)
        # Validar cédula única excluyendo la actual
        cedula = request.POST.get('cedula').strip()
        if Personas.objects.filter(per_cedula=cedula).exclude(pk=persona.pk).exists():
            formPersona.add_error('cedula', 'Ya existe un usuario con este número de cédula.')
        # Validar formularios completos
        if not formUsuario.is_valid() or not formPersona.is_valid():
            messages.error(request, 'Hay errores en el formulario. Por favor revisa los campos resaltados.')
            print("Hay errores en el formulario. Por favor revisa los campos resaltados.: ",formUsuario.errors)
            return render(request,'docente/Edit.html',{
                    'formUsuario': formUsuario,
                'formPersona': formPersona,
                'docente': docente,
                'persona': persona,
            })
        try:
            # Actualizar usuario
            usuario = formUsuario.save(commit=False)
            usuario.first_name = usuario.first_name.title()
            usuario.last_name = usuario.last_name.title()
            nueva_contrasena = request.POST.get('password')
            if nueva_contrasena:
                usuario.set_password(nueva_contrasena)
            usuario.save()
            # Obtener los datos para tabla Personas
            persona.per_segundo_nombre = formPersona.cleaned_data['per_segundo_nombre'].title()
            persona.per_segundo_apellido = formPersona.cleaned_data['per_segundo_apellido'].title()
            persona.per_fecha_nacimiento = formPersona.cleaned_data['per_fecha_nacimiento']
            persona.per_cedula = cedula
            persona.per_telefono = formPersona.cleaned_data['per_telefono']
            persona.save()
            # Obtener los datos para tabla Docentes
            if 'fotografia' in request.FILES:
                docente.doc_fotografia = request.FILES['fotografia']
            docente.doc_estado = True if request.POST.get('estado') == 'True' else False
            docente.save()  # Actualizar los datos de Docentes
            messages.success(request, 'Docente actualizado exitosamente.')
            return redirect('docente_index')
        except Exception as e:
            messages.error(request, f'Error al actualizar el docente: {str(e)}')
        return render(request, 'docente/Edit.html', {
            'formUsuario': formUsuario,
            'formPersona': formPersona,
            'docente': docente,
            'persona': persona,
        })
        
# METODO PARA ELIMINAR UN DOCENTE
@admin_required
def eliminar_docente(request, id_docen):
    docente = get_object_or_404(Docentes, pk=id_docen)  # Obtener el docente por su ID
    persona = docente.fk_id_persona # Obtener la persona asociada al docente
    usuario = persona.fk_id_usuario # Obtener el usuario asociado a la persona
    try:
        usuario.delete()  # Eliminar usuario
    except Exception as e:
        messages.error(request, f'Error al eliminar: {str(e)}')
    
    return redirect('docente_index')
