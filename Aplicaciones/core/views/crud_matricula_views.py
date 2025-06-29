from django.db import IntegrityError, transaction
import pandas as pd
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from ..models import Matriculas, Estudiantes, Clases,Personas, Matriculas,User
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import Font
from django.db.models import Count
from django.contrib.auth.models import User, Group
from django.contrib.auth.decorators import login_required
from ..decorators import admin_required

@admin_required 
def index(request):
    clases = Clases.objects.annotate(
        num_matriculados=Count('matriculas'),
        
    )

    return render(request, 'matricula/Index.html', {'clases': clases})


@admin_required
def detalle(request, cla_id):
    clase = get_object_or_404(Clases, pk=cla_id)
    matriculas = Matriculas.objects.filter(fk_clase=cla_id)
    estudiantes = Estudiantes.objects.all()
    return render(request, 'matricula/Detalle.html', {'matriculas': matriculas,'clase':clase, 'estudiantes': estudiantes})


@admin_required
def matriculaIndividual(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        cla_id = request.POST.get('cla_id')
        est_id = request.POST.get('estudiante')

        clase = get_object_or_404(Clases, pk=cla_id)
        estudiante = get_object_or_404(Estudiantes, pk=est_id)

        if Matriculas.objects.filter(fk_clase=clase, fk_estudiante=estudiante).exists():
            return JsonResponse({'success': False, 'message': 'Estudiante ya está matriculado'})

        Matriculas.objects.create(fk_clase=clase, fk_estudiante=estudiante)

        return JsonResponse({'success': True, 'message': 'Estudiante matriculado exitosamente'})

    return JsonResponse({'success': False, 'message': 'Solicitud inválida'})
    
@admin_required
def vista_tabla_matriculados(request, cla_id):
    matriculas = Matriculas.objects.filter(fk_clase=cla_id).order_by(
        'fk_estudiante__fk_id_persona__fk_id_usuario__last_name',
        'fk_estudiante__fk_id_persona__per_segundo_apellido'
    )
    return render(request, 'matricula/Tabla_Matriculados.html', {'matriculas': matriculas})


@require_POST
@admin_required
def eliminar_matricula(request, matricula_id):
    try:
        matricula = Matriculas.objects.get(pk=matricula_id)
        matricula.delete()
        return JsonResponse({'success': True, 'message': 'Matrícula eliminada correctamente'})
    except Matriculas.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Matrícula no encontrada'})


@csrf_exempt
@admin_required
def importar_estudiantes_excel(request):
    # Verificar método y archivo
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'})
    
    if not request.FILES.get('archivo'):
        return JsonResponse({'success': False, 'message': 'No se ha seleccionado ningún archivo'})
    
    try:
        archivo_excel = request.FILES['archivo']
        cla_id = request.POST.get('cla_id')
        
        # Verificar que cla_id existe
        if not cla_id:
            return JsonResponse({'success': False, 'message': 'ID de clase no proporcionado'})
        
        clase = get_object_or_404(Clases, pk=cla_id)
        
        # Verificar que el archivo es Excel
        if not archivo_excel.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'success': False, 'message': 'El archivo debe ser un Excel (.xlsx o .xls)'})
        
        # Leer el archivo Excel
        try:
            df = pd.read_excel(archivo_excel, dtype=str)
            
        except Exception as e:
            
            return JsonResponse({'success': False, 'message': f'Error al leer el archivo Excel: {str(e)}'})
        
        # Verificar que el DataFrame no esté vacío
        if df.empty:
            return JsonResponse({'success': False, 'message': 'El archivo Excel está vacío'})
        
        # Verificar columnas requeridas
        columnas_requeridas = ['cedula', 'nombre', 'apellido']
        columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
        if columnas_faltantes:
            return JsonResponse({
                'success': False, 
                'message': f'Faltan las siguientes columnas en el Excel: {", ".join(columnas_faltantes)}'
            })
        
        duplicados = []
        agregados = 0
        matriculados_existentes = 0
        
        for index, row in df.iterrows():
            try:
                # Procesar datos de la fila
                nombre = str(row.get('nombre', '')).strip().upper()
                apellido = str(row.get('apellido', '')).strip().upper()
                segundo_nombre = str(row.get('segundo_nombre', '')).strip().upper()
                segundo_apellido = str(row.get('segundo_apellido', '')).strip().upper()
                cedula = str(row.get('cedula', '')).strip().replace('.', '').replace(',', '')
                telefono = str(row.get('telefono', '')).strip().replace('.', '').replace(',', '')
                email = str(row.get('email', f"{cedula}@example.com")).strip()
                
                
                
                # Validar datos mínimos
                if not nombre or not apellido:
                    duplicados.append(f"Fila {index + 2} - Nombre y apellido son obligatorios")
                    continue
                
                # Validar cédula
                if len(cedula) < 5 or not cedula.isdigit():
                    duplicados.append(f"Fila {index + 2} - Cédula inválida: {cedula}")
                    continue
                
                # Usar transacción para cada fila
                try:
                    with transaction.atomic():
                        # Verificar si el usuario ya existe
                        usuario_existente = User.objects.filter(username=cedula).first()
                        
                        if usuario_existente:
                            # El usuario ya existe, buscar si tiene persona y estudiante
                            try:
                                persona = Personas.objects.get(fk_id_usuario=usuario_existente)
                                estudiante = Estudiantes.objects.get(fk_id_persona=persona)
                                
                                # Verificar si ya está matriculado en esta clase
                                if Matriculas.objects.filter(fk_clase=clase, fk_estudiante=estudiante).exists():
                                    duplicados.append(f"Fila {index + 2} - Estudiante ya matriculado en esta clase")
                                    continue
                                else:
                                    # Matricular estudiante existente
                                    Matriculas.objects.create(fk_clase=clase, fk_estudiante=estudiante)
                                    matriculados_existentes += 1
                                    
                                    continue
                                    
                            except (Personas.DoesNotExist, Estudiantes.DoesNotExist):
                                duplicados.append(f"Fila {index + 2} - Usuario existe pero no tiene perfil de estudiante completo")
                                continue
                        
                        # Verificar si existe persona con esta cédula pero sin usuario
                        persona_existente = Personas.objects.filter(per_cedula=cedula).first()
                        
                        if persona_existente and persona_existente.fk_id_usuario is None:
                            duplicados.append(f"Fila {index + 2} - Existe persona con esta cédula sin usuario asociado")
                            continue
                        elif persona_existente and persona_existente.fk_id_usuario:
                            duplicados.append(f"Fila {index + 2} - Persona ya existe con usuario")
                            continue
                        
                        # Crear nuevo usuario, persona y estudiante
                        # Crear usuario
                        usuario = User.objects.create_user(
                            username=cedula,
                            first_name=nombre,
                            last_name=apellido,
                            email=email,
                            password=cedula
                        )
                        
                        # Asignar al grupo de estudiantes
                        grupo_estudiantes, _ = Group.objects.get_or_create(name="Estudiantes")
                        usuario.groups.add(grupo_estudiantes)
                        
                        # Crear persona
                        persona = Personas.objects.create(
                            fk_id_usuario=usuario,
                            per_segundo_nombre=segundo_nombre,
                            per_segundo_apellido=segundo_apellido,
                            per_cedula=cedula,
                            per_telefono=telefono,
                            per_fecha_nacimiento=None
                        )
                        
                        # Crear estudiante
                        estudiante = Estudiantes.objects.create(fk_id_persona=persona)
                        
                        # Matricular en la clase
                        Matriculas.objects.create(fk_clase=clase, fk_estudiante=estudiante)
                        
                        agregados += 1
                        
                        
                except IntegrityError as ie:
                    # Error de integridad (usuario duplicado, etc.)
                    duplicados.append(f"Fila {index + 2} - Error de integridad: Usuario o cédula ya existe")
                    
                    continue
            
            except Exception as e:
                duplicados.append(f"Fila {index + 2} - Error inesperado: {str(e)}")
                
                continue
        
        # Preparar mensaje de respuesta
        mensaje = f"{agregados} estudiante(s) nuevo(s) importado(s).\n{matriculados_existentes} estudiante(s) existentes fueron matriculados."
        
        if duplicados:
            mensaje += f"\n{len(duplicados)} advertencia(s):\n" + "\n".join(duplicados)
        
        
        
        return JsonResponse({'success': True, 'message': mensaje})
    
    except Exception as e:
        
        return JsonResponse({'success': False, 'message': f'Error interno del servidor: {str(e)}'})
    
    # Esta línea nunca debería ejecutarse, pero por seguridad
    return JsonResponse({'success': False, 'message': 'Solicitud inválida'})



def descargar_plantilla_estudiantes(request):
    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    # Encabezados ordenados y estilizados
    columnas = ['nombre', 'segundo_nombre', 'apellido', 'segundo_apellido', 'cedula', 'telefono', 'email']
    ws.append(columnas)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Fila de ejemplo (valores como texto, conservando ceros)
    ejemplo = ['Juan', 'Carlos', 'Pérez', 'González', '0991234567', '0912345678', 'juan.perez@example.com']
    ws.append(ejemplo)

    # Aplicar formato texto explícito a cédula y teléfono (columnas 5 y 6: E y F)
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = '@'

    # Ajuste de anchos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # Preparar archivo para descarga
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_estudiantes.xlsx'
    return response
@admin_required
def importar(request):
    clases = Clases.objects.all()
    resultados = {}
    advertencias = []

    # Descargar plantilla
    if request.method == 'POST' and 'descargar_plantilla' in request.POST:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Estudiantes'
        ws.append([
            'Nombres', 'Apellidos', 'Segundo Nombre', 'Segundo Apellido',
            'Cédula', 'Teléfono', 'Fecha Nacimiento',
            'Contacto Emergencia', 'Teléfono Emergencia'
        ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=plantilla_estudiantes.xlsx'
        wb.save(response)
        return response

    # Importar estudiantes
    elif request.method == 'POST' and 'importar_excel' in request.POST:
        archivo_excel = request.FILES.get('archivo_excel')
        clase_id = request.POST.get('fk_clase')

        if archivo_excel and clase_id:
            wb = load_workbook(filename=archivo_excel)
            ws = wb.active
            clase = get_object_or_404(Clases, pk=clase_id)
            agregados = 0
            matriculados_existentes = 0

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                nombre, apellido, segundo_nombre, segundo_apellido, cedula, telefono, fecha_nac, contacto_emer, tel_emer = row

                if not cedula or not str(cedula).isdigit() or len(str(cedula)) < 5:
                    advertencias.append(f"Fila {i+2}: Cédula inválida.")
                    continue

                cedula = str(cedula).strip()
                nombre = str(nombre or '').strip().title()
                apellido = str(apellido or '').strip().title()
                segundo_nombre = str(segundo_nombre or '').strip().title()
                segundo_apellido = str(segundo_apellido or '').strip().title()
                telefono = str(telefono or '').strip()
                email = f"{cedula}@example.com"
                
                print(f"Nombre procesado: {nombre}, Apellido: {apellido}")
                # Verificar si ya existe el usuario
                usuario = User.objects.filter(username=cedula).first()

                if usuario:
                    # Verificar si existe persona asociada
                    persona = Personas.objects.filter(fk_id_usuario=usuario).first()
                    if not persona:
                        persona = Personas.objects.create(
                            fk_id_usuario=usuario,
                            per_segundo_nombre=segundo_nombre,
                            per_segundo_apellido=segundo_apellido,
                            per_cedula=cedula,
                            per_telefono=telefono,
                            per_fecha_nacimiento=fecha_nac
                        )

                    # Verificar si existe estudiante asociado
                    estudiante = Estudiantes.objects.filter(fk_id_persona=persona).first()
                    if not estudiante:
                        estudiante = Estudiantes.objects.create(
                            fk_id_persona=persona,
                            est_contacto_emergencia=contacto_emer,
                            est_telefono_emergencia=tel_emer
                        )

                    # Verificar si ya está matriculado
                    if Matriculas.objects.filter(fk_clase=clase, fk_estudiante=estudiante).exists():
                        advertencias.append(f"Fila {i+2}: Estudiante ya matriculado.")
                    else:
                        Matriculas.objects.create(fk_clase=clase, fk_estudiante=estudiante)
                        matriculados_existentes += 1
                    continue  # Pasar a la siguiente fila

                # Si no existe el usuario, lo creamos
                usuario = User.objects.create_user(
                    username=cedula,
                    first_name=nombre,
                    last_name=apellido,
                    email=email,
                    password=cedula
                )

                grupo_estudiantes, _ = Group.objects.get_or_create(name="Estudiantes")
                usuario.groups.add(grupo_estudiantes)

                persona = Personas.objects.create(
                    fk_id_usuario=usuario,
                    per_segundo_nombre=segundo_nombre,
                    per_segundo_apellido=segundo_apellido,
                    per_cedula=cedula,
                    per_telefono=telefono,
                    per_fecha_nacimiento=fecha_nac
                )

                estudiante = Estudiantes.objects.create(
                    fk_id_persona=persona,
                    est_contacto_emergencia=contacto_emer,
                    est_telefono_emergencia=tel_emer
                )

                Matriculas.objects.create(fk_clase=clase, fk_estudiante=estudiante)
                agregados += 1

            mensaje = f"{agregados} estudiante(s) nuevo(s) importado(s). {matriculados_existentes} estudiante(s) existente(s) matriculado(s)."
            if advertencias:
                mensaje += f"\n{len(advertencias)} advertencia(s):\n" + "\n".join(advertencias)

            messages.success(request, mensaje)
        else:
            messages.error(request, "Debe seleccionar una clase y subir un archivo.")

    return render(request, 'matricula/Importar.html', {
        'clases': clases,
        'resultados': resultados
    })